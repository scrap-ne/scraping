from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font


COLUNAS = ["Estado", "link", "tema", "data", "titulo", "descrição"]


def exportar_excel(resultados, caminho):
    caminho = Path(caminho).resolve()
    caminho.parent.mkdir(parents=True, exist_ok=True)

    tabela = pd.DataFrame(resultados, columns=COLUNAS)
    tabela = tabela.drop_duplicates(subset=COLUNAS, keep="first")

    datas = pd.to_datetime(tabela["data"], errors="coerce")
    tabela["_data_ordenacao"] = datas
    tabela = tabela.sort_values(
        ["Estado", "_data_ordenacao", "tema", "titulo"],
        ascending=[True, False, True, True],
        na_position="last",
    )
    tabela["data"] = datas.loc[tabela.index].apply(
        lambda valor: valor.date() if pd.notna(valor) else None
    )
    tabela = tabela.drop(columns="_data_ordenacao")

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        tabela.to_excel(writer, index=False, sheet_name="Resultados")
        planilha = writer.sheets["Resultados"]
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = planilha.dimensions

        larguras = {
            "A": 20,
            "B": 70,
            "C": 42,
            "D": 13,
            "E": 55,
            "F": 100,
        }
        for coluna, largura in larguras.items():
            planilha.column_dimensions[coluna].width = largura

        for celula in planilha[1]:
            celula.font = Font(bold=True)
            celula.alignment = Alignment(horizontal="center")

        for linha in range(2, planilha.max_row + 1):
            planilha.cell(linha, 4).number_format = "DD/MM/YYYY"
            planilha.cell(linha, 5).alignment = Alignment(vertical="top", wrap_text=True)
            planilha.cell(linha, 6).alignment = Alignment(vertical="top", wrap_text=True)

            celula_link = planilha.cell(linha, 2)
            if celula_link.value:
                celula_link.hyperlink = celula_link.value
                celula_link.style = "Hyperlink"

    return caminho, len(tabela)
